"""
Rozšířené zadání
Modul openpyxl lze používat i samostatně. Přehled jeho funkcí najdeš v dokumentaci.
Zkus pomocí modulu zapsat rozvrh hodin jako tabulku v Excel sešitu. Níže máš program, který obsahuje
rozvrh hodin jako dvourozměrný seznam. Vnitřní seznamy obsahují předměty v rozvrhu,
jeden vnitřní seznam vždy obsahuje předměty pro daný den.
Podívej se nejprve na ukázku. Jednoduchý program níže zapíše hodnotu Test do buňky B1 a
výsledek uloží souboru rozvrh_hodin.xlsx. Pokud neznáš terminologii Excelu,
pak Workbook označuje sešit, tj. celý "soubor". ws se odkazuje na Work Sheet, což je list,
tj. jedna "záložka". Náš soubor bude mít jen jeden list, čímž je situace jednoduchá.

from openpyxl import Workbook
wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"
cislo_radku = 1
cislo_sloupce = 2
bunka = ws1.cell(cislo_radku, cislo_sloupce)
bunka.value = "Test"
wb.save(filename="rozvrh_hodin.xlsx")

Dobrovolný doplněk rozšířeného zadání :-)
Modul openpyxl je silný v možnostech formátování, které nabízí. Podívej se na rozšířený příklad níže.
Ten kromě zápisu hodnoty nastaví buňce šedivou barvu jako pozadí.
Přehled barev, které můžeš použít, najdeš v dokumentaci.

from openpyxl import Workbook
from openpyxl.styles import PatternFill
wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"
cislo_radku = 1
cislo_sloupce = 2

bunka = ws1.cell(cislo_radku, cislo_sloupce)
bunka.value = "Test"

# Určím si barvu
sediva_barva = PatternFill("solid", fgColor="00C0C0C0")
# Přiřadím barvu buňce
bunka.fill = sediva_barva

wb.save(filename="rozvrh_hodin.xlsx")
Vrať se ke svému programu a zkus vytvořit rozvrh obarvený. Například můžeš zkusit podbarvit oběd šedě
a poté třeba nastavit nějakou barvu výchovám, jinou barvu jazykům atd.
"""
import pandas
import openpyxl

from openpyxl import Workbook
from openpyxl.styles import PatternFill
wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"

rozvrh_hodin = [
  ["Anglický jazyk", "Přírodopis", "Dějepis", "Matematika", "Oběd", "Tělesná výchova", "Tělesná výchova", ],
  ["Občanská výchova", "Hudební výchova", "Matematika", "Oběd", "Výtvarná výchova", "Dějepis", ],
  ["Matematika", "Chemie", "Přírodopis", "Fyzika", "Oběd", "Zeměpis", ],
  ["Fyzika", "Anglický jazyk", "Matematika", "Český jazyk", "Dějepis", "Oběd", ],
  ["Český jazyk", "Zeměpis", "Český jazyk", "Výtvarná výchova", "Oběd", ]
]

grey = PatternFill("solid", fgColor="00C0C0C0")
green = PatternFill("solid", fgColor="0000FF00")
red = PatternFill("solid", fgColor="00FF0000")

radek = 1
for den in rozvrh_hodin:
  sloupec = 1
  for predmet in den:
    bunka = ws1.cell(radek,sloupec)
    bunka.value = predmet
    if bunka.value == "Oběd":
      bunka.fill = grey
    elif bunka.value == "Tělesná výchova" or bunka.value == "Výtvarná výchova" or bunka.value == "Hudební výchova" or bunka.value =="Občanská výchova":
      bunka.fill = green
    else:
      bunka.fill = red
    sloupec += 1
  radek += 1
wb.save(filename="rozvrh_hodin.xlsx")